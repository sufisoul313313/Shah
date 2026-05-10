"""
Canada PR Application Status Tracker
Reads PR_Tracker.xlsx and generates a formatted status report.
"""

import sys
import os
from datetime import datetime, timedelta
from pathlib import Path

try:
    import pandas as pd
    from openpyxl import load_workbook
    from openpyxl import Workbook
    from openpyxl.styles import (
        PatternFill, Font, Alignment, Border, Side
    )
    from openpyxl.utils import get_column_letter
except ImportError:
    print("Missing dependencies. Run: pip install pandas openpyxl")
    sys.exit(1)

EXCEL_PATH = Path.home() / "Downloads" / "PR_Tracker.xlsx"
REPORTS_DIR = Path(__file__).parent / "reports"
TODAY = datetime.today().date()
EXPIRY_ALERT_DAYS = 30

REQUIRED_COLUMNS = {
    "document name", "status", "date submitted", "expiry date", "notes"
}


def load_tracker(path: Path) -> pd.DataFrame:
    if not path.exists():
        print(f"ERROR: Excel file not found at: {path}")
        print("Creating a sample file there so you can fill it in.")
        create_sample_excel(path)
        sys.exit(0)

    df = pd.read_excel(path, engine="openpyxl")
    df.columns = df.columns.str.strip().str.lower()

    missing = REQUIRED_COLUMNS - set(df.columns)
    if missing:
        print(f"ERROR: Missing columns in Excel: {missing}")
        print(f"Found columns: {list(df.columns)}")
        sys.exit(1)

    # Normalize dates
    for col in ("date submitted", "expiry date"):
        df[col] = pd.to_datetime(df[col], errors="coerce").dt.date

    df["status"] = df["status"].str.strip().str.title()
    df["document name"] = df["document name"].str.strip()
    df["notes"] = df["notes"].fillna("").str.strip()

    return df


def create_sample_excel(path: Path):
    path.parent.mkdir(parents=True, exist_ok=True)
    import openpyxl
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "PR Tracker"

    headers = ["Document Name", "Status", "Date Submitted", "Expiry Date", "Notes"]
    ws.append(headers)

    sample_rows = [
        ["Medical Exam",          "Submitted", "2025-11-01", "2026-11-01", "Done at LifeLabs"],
        ["Police Certificate",    "Approved",  "2025-10-15", "2026-10-15", ""],
        ["Birth Certificate",     "Pending",   "",           "",           "Need certified copy"],
        ["Passport Copy",         "Approved",  "2025-09-01", "",           ""],
        ["Proof of Funds",        "Submitted", "2025-12-01", "",           "Bank statement"],
        ["Language Test (IELTS)", "Approved",  "2024-06-01", "2026-06-01", "Score: 8.0"],
        ["ECA Report",            "Expired",   "2023-01-10", "2025-01-10", "Needs renewal"],
        ["Photographs",           "Pending",   "",           "",           ""],
    ]
    for row in sample_rows:
        ws.append(row)

    # Auto-width columns
    for col in ws.columns:
        max_len = max(len(str(cell.value or "")) for cell in col) + 4
        ws.column_dimensions[col[0].column_letter].width = max_len

    wb.save(path)
    print(f"Sample file created: {path}")
    print("Fill it in with your real data and run this script again.")


# ── colour palette ───────────────────────────────────────────────
FILL = {
    "header":    PatternFill("solid", fgColor="1F4E79"),
    "section":   PatternFill("solid", fgColor="2E75B6"),
    "Approved":  PatternFill("solid", fgColor="C6EFCE"),
    "Submitted": PatternFill("solid", fgColor="FFEB9C"),
    "Pending":   PatternFill("solid", fgColor="FCE4D6"),
    "Expired":   PatternFill("solid", fgColor="F4CCCC"),
    "alert":     PatternFill("solid", fgColor="FF0000"),
    "summary":   PatternFill("solid", fgColor="DEEAF1"),
    "steps":     PatternFill("solid", fgColor="E2EFDA"),
}
FONT_WHITE_BOLD = Font(bold=True, color="FFFFFF", size=11)
FONT_BOLD       = Font(bold=True, size=10)
FONT_NORMAL     = Font(size=10)
FONT_RED_BOLD   = Font(bold=True, color="CC0000", size=10)
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT   = Alignment(horizontal="left",   vertical="center", wrap_text=True)

def _thin_border():
    s = Side(style="thin", color="BFBFBF")
    return Border(left=s, right=s, top=s, bottom=s)


def _expiring_soon_mask(df: pd.DataFrame) -> pd.Series:
    def _check(d):
        if d is None or d != d:
            return False
        try:
            return TODAY <= d <= TODAY + timedelta(days=EXPIRY_ALERT_DAYS)
        except TypeError:
            return False
    return df["expiry date"].apply(_check)


def _set_col_widths(ws, widths: list[int]):
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


def _section_row(ws, text: str, ncols: int):
    r = ws.max_row + 1
    ws.append([""] * ncols)
    ws.cell(r, 1).value = text
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=ncols)
    cell = ws.cell(r, 1)
    cell.fill = FILL["section"]
    cell.font = FONT_WHITE_BOLD
    cell.alignment = LEFT
    ws.row_dimensions[r].height = 18


def _header_row(ws, labels: list[str]):
    ws.append(labels)
    r = ws.max_row
    for c, _ in enumerate(labels, start=1):
        cell = ws.cell(r, c)
        cell.fill = FILL["header"]
        cell.font = FONT_WHITE_BOLD
        cell.alignment = CENTER
        cell.border = _thin_border()
    ws.row_dimensions[r].height = 20


def save_report(df: pd.DataFrame) -> Path:
    REPORTS_DIR.mkdir(exist_ok=True)
    out = REPORTS_DIR / f"PR_Status_Report_{TODAY.strftime('%Y-%m-%d')}.xlsx"

    total     = len(df)
    approved  = df[df["status"] == "Approved"]
    submitted = df[df["status"] == "Submitted"]
    pending   = df[df["status"] == "Pending"]
    expired   = df[df["status"] == "Expired"]
    pct       = round(len(approved) / total * 100) if total else 0
    soon_mask = _expiring_soon_mask(df)
    expiring  = df[soon_mask]

    wb = Workbook()

    # ── Sheet 1: Summary ─────────────────────────────────────────
    ws = wb.active
    ws.title = "Summary"
    _set_col_widths(ws, [28, 16, 16, 16, 40])
    ws.row_dimensions[1].height = 30

    # Title
    ws.append(["CANADA PR APPLICATION STATUS REPORT", "", "", "", ""])
    ws.merge_cells("A1:E1")
    ws["A1"].fill   = FILL["header"]
    ws["A1"].font   = Font(bold=True, color="FFFFFF", size=14)
    ws["A1"].alignment = CENTER

    ws.append([f"Generated: {TODAY.strftime('%B %d, %Y')}", "", "", "", ""])
    ws.merge_cells("A2:E2")
    ws["A2"].font = Font(italic=True, size=10)
    ws["A2"].alignment = CENTER

    ws.append([])  # spacer

    # Progress summary block
    _section_row(ws, "OVERALL PROGRESS", 5)
    for label, val in [
        ("Total Documents", total),
        ("Approved",        len(approved)),
        ("Submitted",       len(submitted)),
        ("Pending",         len(pending)),
        ("Expired",         len(expired)),
        ("% Complete",      f"{pct}%"),
    ]:
        r = ws.max_row + 1
        ws.append([label, val, "", "", ""])
        ws.cell(r, 1).font = FONT_BOLD
        ws.cell(r, 1).fill = FILL["summary"]
        ws.cell(r, 2).fill = FILL["summary"]
        ws.cell(r, 2).alignment = CENTER

    ws.append([])  # spacer

    # Expiring soon alert
    _section_row(ws, "⚠  EXPIRING WITHIN 30 DAYS", 5)
    if expiring.empty:
        r = ws.max_row + 1
        ws.append(["No documents expiring soon.", "", "", "", ""])
        ws.cell(r, 1).font = FONT_NORMAL
    else:
        _header_row(ws, ["Document", "Status", "Expiry Date", "Days Left", "Notes"])
        for _, row in expiring.iterrows():
            days_left = (row["expiry date"] - TODAY).days
            r = ws.max_row + 1
            ws.append([
                row["document name"],
                row["status"],
                str(row["expiry date"]),
                days_left,
                row["notes"],
            ])
            for c in range(1, 6):
                ws.cell(r, c).fill   = FILL["alert"] if days_left <= 7 else PatternFill("solid", fgColor="FFEB9C")
                ws.cell(r, c).font   = FONT_RED_BOLD if days_left <= 7 else FONT_NORMAL
                ws.cell(r, c).border = _thin_border()

    ws.append([])

    # Next steps
    _section_row(ws, "NEXT STEPS", 5)
    steps = []
    for _, row in pending.iterrows():
        steps.append(f"[ ]  Obtain and submit: {row['document name']}")
    for _, row in expiring.iterrows():
        steps.append(f"[ ]  Renew before expiry: {row['document name']}")
    for _, row in expired.iterrows():
        steps.append(f"[ ]  Renew expired document: {row['document name']}")
    if not steps:
        steps = ["All documents are in order. Monitor for upcoming expiries."]
    for step in steps:
        r = ws.max_row + 1
        ws.append([step, "", "", "", ""])
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=5)
        ws.cell(r, 1).fill      = FILL["steps"]
        ws.cell(r, 1).font      = FONT_NORMAL
        ws.cell(r, 1).alignment = LEFT

    # ── Sheet 2: All Documents ───────────────────────────────────
    ws2 = wb.create_sheet("All Documents")
    _set_col_widths(ws2, [30, 14, 16, 16, 40])

    ws2.append(["FULL DOCUMENT LIST", "", "", "", ""])
    ws2.merge_cells("A1:E1")
    ws2["A1"].fill      = FILL["header"]
    ws2["A1"].font      = FONT_WHITE_BOLD
    ws2["A1"].alignment = CENTER
    ws2.row_dimensions[1].height = 24

    ws2.append([])
    _header_row(ws2, ["Document Name", "Status", "Date Submitted", "Expiry Date", "Notes"])

    status_colors = {k: FILL[k] for k in ("Approved", "Submitted", "Pending", "Expired")}

    for _, row in df.iterrows():
        sub = str(row["date submitted"]) if pd.notna(row["date submitted"]) else "—"
        exp = str(row["expiry date"])    if pd.notna(row["expiry date"])    else "—"
        r = ws2.max_row + 1
        ws2.append([row["document name"], row["status"], sub, exp, row["notes"]])
        fill = status_colors.get(row["status"], PatternFill())
        for c in range(1, 6):
            ws2.cell(r, c).fill      = fill
            ws2.cell(r, c).font      = FONT_NORMAL
            ws2.cell(r, c).border    = _thin_border()
            ws2.cell(r, c).alignment = LEFT
        # Highlight entire row red if expired
        if row["status"] == "Expired":
            for c in range(1, 6):
                ws2.cell(r, c).font = FONT_RED_BOLD

    # freeze header rows
    ws2.freeze_panes = "A4"

    wb.save(out)
    return out


def main():
    excel_path = Path(sys.argv[1]) if len(sys.argv) > 1 else EXCEL_PATH

    print(f"Loading: {excel_path}")
    df = load_tracker(excel_path)
    print(f"Loaded {len(df)} documents.")

    out_path = save_report(df)
    print(f"Report saved to: {out_path}")


if __name__ == "__main__":
    main()
